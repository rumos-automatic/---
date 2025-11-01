#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excelテンプレート作成スクリプト
会社事業における融資と利益の理解 - 付録D用
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 出力先ディレクトリ
OUTPUT_DIR = r"C:\CLAUDE_CODE\01_RUMOS_Project\会社事業\assets\templates"

# 共通スタイル定義
TITLE_FONT = Font(name='メイリオ', size=16, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
HEADER_FONT = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
INPUT_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CALC_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
TOTAL_FONT = Font(name='メイリオ', size=11, bold=True)
TOTAL_FILL = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
NORMAL_FONT = Font(name='メイリオ', size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def set_column_width(ws, col, width):
    """列幅を設定"""
    ws.column_dimensions[get_column_letter(col)].width = width


def create_template_1():
    """1. 月次損益計算書"""
    wb = Workbook()
    ws = wb.active
    ws.title = "月次損益計算書"

    # タイトル
    ws.merge_cells('A1:D1')
    ws['A1'] = "月次損益計算書"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:D2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 対象月
    ws['A4'] = "対象月"
    ws['A4'].font = HEADER_FONT
    ws['A4'].fill = HEADER_FILL
    ws['A4'].alignment = CENTER_ALIGN
    ws['B4'] = "2025年1月"
    ws['B4'].fill = INPUT_FILL
    ws['B4'].alignment = CENTER_ALIGN

    # ヘッダー
    ws['A6'] = "科目"
    ws['B6'] = "金額（円）"
    ws['C6'] = "構成比（%）"
    ws['D6'] = "備考"
    for col in ['A6', 'B6', 'C6', 'D6']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN
        ws[col].border = THIN_BORDER

    # 売上
    row = 7
    ws[f'A{row}'] = "売上高"
    ws[f'A{row}'].font = NORMAL_FONT
    ws[f'B{row}'] = 5000000
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = f'=B{row}/$B${row}*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = "売上の合計"

    # 売上原価
    row = 8
    ws[f'A{row}'] = "売上原価"
    ws[f'B{row}'] = 2000000
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = f'=B{row}/$B$7*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = "商品の仕入・製造原価"

    # 売上総利益
    row = 9
    ws[f'A{row}'] = "売上総利益（粗利）"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = "=B7-B8"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    ws[f'C{row}'] = f'=B{row}/$B$7*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = "売上高-売上原価"

    # 販管費
    row = 11
    ws[f'A{row}'] = "販売費及び一般管理費"
    ws[f'A{row}'].font = HEADER_FONT

    expenses = [
        ("人件費", 800000, "給与・賞与"),
        ("広告宣伝費", 300000, "マーケティング費用"),
        ("家賃", 200000, "事務所・店舗家賃"),
        ("水道光熱費", 50000, "電気・水道・ガス"),
        ("通信費", 30000, "電話・インターネット"),
        ("その他経費", 120000, "消耗品・雑費など")
    ]

    start_row = 12
    for i, (name, amount, note) in enumerate(expenses):
        r = start_row + i
        ws[f'A{r}'] = f"  {name}"
        ws[f'B{r}'] = amount
        ws[f'B{r}'].fill = INPUT_FILL
        ws[f'B{r}'].number_format = '#,##0'
        ws[f'C{r}'] = f'=B{r}/$B$7*100'
        ws[f'C{r}'].fill = CALC_FILL
        ws[f'C{r}'].number_format = '0.0'
        ws[f'D{r}'] = note

    # 販管費合計
    row = start_row + len(expenses)
    ws[f'A{row}'] = "販管費合計"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = f"=SUM(B{start_row}:B{row-1})"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    ws[f'C{row}'] = f'=B{row}/$B$7*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'

    # 営業利益
    row += 2
    ws[f'A{row}'] = "営業利益"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = "=B9-B18"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    ws[f'C{row}'] = f'=B{row}/$B$7*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = "本業の利益"

    # 営業外収益
    row += 1
    ws[f'A{row}'] = "営業外収益"
    ws[f'B{row}'] = 10000
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'D{row}'] = "受取利息など"

    # 営業外費用
    row += 1
    ws[f'A{row}'] = "営業外費用"
    ws[f'B{row}'] = 50000
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'D{row}'] = "支払利息など"

    # 経常利益
    row += 1
    ws[f'A{row}'] = "経常利益"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = "=B20+B21-B22"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    ws[f'C{row}'] = f'=B{row}/$B$7*100'
    ws[f'C{row}'].fill = CALC_FILL
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = "営業利益±営業外損益"

    # 列幅設定
    set_column_width(ws, 1, 25)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 15)
    set_column_width(ws, 4, 30)

    wb.save(os.path.join(OUTPUT_DIR, "月次損益計算書.xlsx"))
    print("[OK] 月次損益計算書.xlsx 作成完了")


def create_template_2():
    """2. 貸借対照表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "貸借対照表"

    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = "貸借対照表（Balance Sheet）"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:F2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 基準日
    ws['A4'] = "基準日"
    ws['A4'].font = HEADER_FONT
    ws['A4'].fill = HEADER_FILL
    ws['B4'] = "2025年1月31日"
    ws['B4'].fill = INPUT_FILL
    ws['B4'].alignment = CENTER_ALIGN

    # ヘッダー（左側：資産）
    ws.merge_cells('A6:C6')
    ws['A6'] = "【資産の部】"
    ws['A6'].font = HEADER_FONT
    ws['A6'].fill = HEADER_FILL
    ws['A6'].alignment = CENTER_ALIGN

    # ヘッダー（右側：負債・純資産）
    ws.merge_cells('D6:F6')
    ws['D6'] = "【負債・純資産の部】"
    ws['D6'].font = HEADER_FONT
    ws['D6'].fill = HEADER_FILL
    ws['D6'].alignment = CENTER_ALIGN

    # 資産の部
    ws['A7'] = "科目"
    ws['B7'] = "金額（円）"
    ws['C7'] = "%"
    for col in ['A7', 'B7', 'C7']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN

    assets = [
        ("流動資産", None),
        ("  現金・預金", 3000000),
        ("  売掛金", 1500000),
        ("  在庫", 800000),
        ("流動資産合計", "=SUM(B9:B11)"),
        ("固定資産", None),
        ("  建物・設備", 2000000),
        ("  車両", 500000),
        ("固定資産合計", "=SUM(B15:B16)"),
        ("資産合計", "=B12+B17"),
    ]

    row = 8
    for name, value in assets:
        ws[f'A{row}'] = name
        if value is not None:
            if isinstance(value, str):
                ws[f'B{row}'] = value
                ws[f'B{row}'].fill = CALC_FILL
                ws[f'B{row}'].font = TOTAL_FONT if "合計" in name else NORMAL_FONT
            else:
                ws[f'B{row}'] = value
                ws[f'B{row}'].fill = INPUT_FILL
            ws[f'B{row}'].number_format = '#,##0'
            if row != 8 and row != 13:
                ws[f'C{row}'] = f'=B{row}/$B$18*100'
                ws[f'C{row}'].fill = CALC_FILL
                ws[f'C{row}'].number_format = '0.0'
        if "合計" in name:
            ws[f'A{row}'].font = TOTAL_FONT
        row += 1

    # 負債・純資産の部
    ws['D7'] = "科目"
    ws['E7'] = "金額（円）"
    ws['F7'] = "%"
    for col in ['D7', 'E7', 'F7']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN

    liabilities = [
        ("流動負債", None),
        ("  買掛金", 600000),
        ("  短期借入金", 1000000),
        ("流動負債合計", "=SUM(E9:E10)"),
        ("固定負債", None),
        ("  長期借入金", 2000000),
        ("固定負債合計", "=E13"),
        ("負債合計", "=E11+E14"),
        ("純資産", None),
        ("  資本金", 2000000),
        ("  利益剰余金", 2200000),
        ("純資産合計", "=SUM(E17:E18)"),
        ("負債・純資産合計", "=E15+E19"),
    ]

    row = 8
    for name, value in liabilities:
        ws[f'D{row}'] = name
        if value is not None:
            if isinstance(value, str):
                ws[f'E{row}'] = value
                ws[f'E{row}'].fill = CALC_FILL
                ws[f'E{row}'].font = TOTAL_FONT if "合計" in name else NORMAL_FONT
            else:
                ws[f'E{row}'] = value
                ws[f'E{row}'].fill = INPUT_FILL
            ws[f'E{row}'].number_format = '#,##0'
            if row != 8 and row != 12 and row != 16:
                ws[f'F{row}'] = f'=E{row}/$E$20*100'
                ws[f'F{row}'].fill = CALC_FILL
                ws[f'F{row}'].number_format = '0.0'
        if "合計" in name:
            ws[f'D{row}'].font = TOTAL_FONT
        row += 1

    # 財務指標
    ws.merge_cells('A22:F22')
    ws['A22'] = "【主要財務指標】"
    ws['A22'].font = HEADER_FONT
    ws['A22'].fill = HEADER_FILL
    ws['A22'].alignment = CENTER_ALIGN

    ws['A23'] = "自己資本比率"
    ws['B23'] = "=E19/E20*100"
    ws['B23'].fill = CALC_FILL
    ws['B23'].number_format = '0.0"%"'
    ws['C23'] = "（30%以上が健全）"

    ws['A24'] = "流動比率"
    ws['B24'] = "=B12/E11*100"
    ws['B24'].fill = CALC_FILL
    ws['B24'].number_format = '0.0"%"'
    ws['C24'] = "（100%以上が健全）"

    # 列幅設定
    set_column_width(ws, 1, 20)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 10)
    set_column_width(ws, 4, 20)
    set_column_width(ws, 5, 15)
    set_column_width(ws, 6, 10)

    wb.save(os.path.join(OUTPUT_DIR, "貸借対照表.xlsx"))
    print("[OK] 貸借対照表.xlsx 作成完了")


def create_template_3():
    """3. キャッシュフロー計算書"""
    wb = Workbook()
    ws = wb.active
    ws.title = "キャッシュフロー計算書"

    # タイトル
    ws.merge_cells('A1:D1')
    ws['A1'] = "キャッシュフロー計算書（簡易版）"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:D2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 対象期間
    ws['A4'] = "対象期間"
    ws['A4'].font = HEADER_FONT
    ws['A4'].fill = HEADER_FILL
    ws['B4'] = "2025年1月"
    ws['B4'].fill = INPUT_FILL
    ws['B4'].alignment = CENTER_ALIGN

    # ヘッダー
    ws['A6'] = "項目"
    ws['B6'] = "金額（円）"
    ws['C6'] = "説明"
    for col in ['A6', 'B6', 'C6']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN

    # 営業活動によるキャッシュフロー
    ws['A7'] = "【営業活動によるCF】"
    ws['A7'].font = HEADER_FONT

    operating_cf = [
        ("税引前当期純利益", 1500000, "P/Lの利益"),
        ("減価償却費", 100000, "現金流出のない経費"),
        ("売掛金の増減", -200000, "増加はマイナス"),
        ("在庫の増減", -100000, "増加はマイナス"),
        ("買掛金の増減", 150000, "増加はプラス"),
        ("法人税等の支払", -300000, "税金の支払い"),
    ]

    row = 8
    for name, value, note in operating_cf:
        ws[f'A{row}'] = f"  {name}"
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = note
        row += 1

    ws[f'A{row}'] = "営業CF合計"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = f"=SUM(B8:B{row-1})"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    row += 1

    # 投資活動によるキャッシュフロー
    ws[f'A{row}'] = "【投資活動によるCF】"
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    investing_cf = [
        ("固定資産の取得", -500000, "設備投資"),
        ("固定資産の売却", 0, "設備の売却"),
    ]

    for name, value, note in investing_cf:
        ws[f'A{row}'] = f"  {name}"
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = note
        row += 1

    ws[f'A{row}'] = "投資CF合計"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = f"=SUM(B{row-2}:B{row-1})"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    row += 1

    # 財務活動によるキャッシュフロー
    ws[f'A{row}'] = "【財務活動によるCF】"
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    financing_cf = [
        ("借入金の増加", 1000000, "融資の受取"),
        ("借入金の返済", -300000, "融資の返済"),
    ]

    for name, value, note in financing_cf:
        ws[f'A{row}'] = f"  {name}"
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = note
        row += 1

    ws[f'A{row}'] = "財務CF合計"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = f"=SUM(B{row-2}:B{row-1})"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    row += 2

    # 現金の増減
    ws[f'A{row}'] = "現金の増減"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = "=B14+B18+B22"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT
    row += 1

    ws[f'A{row}'] = "期首現金残高"
    ws[f'B{row}'] = 2000000
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'] = "期末現金残高"
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'B{row}'] = f"=B{row-2}+B{row-1}"
    ws[f'B{row}'].fill = TOTAL_FILL
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = TOTAL_FONT

    # 列幅設定
    set_column_width(ws, 1, 30)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 30)

    wb.save(os.path.join(OUTPUT_DIR, "キャッシュフロー計算書.xlsx"))
    print("[OK] キャッシュフロー計算書.xlsx 作成完了")


def create_template_4():
    """4. 資金繰り表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "資金繰り表"

    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = "資金繰り表（3ヶ月版）"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:F2')
    ws['A2'] = "【使い方】青色のセルに予定金額を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # ヘッダー
    ws['A4'] = "項目"
    ws['B4'] = "1月"
    ws['C4'] = "2月"
    ws['D4'] = "3月"
    ws['E4'] = "合計"
    ws['F4'] = "備考"
    for col in ['A4', 'B4', 'C4', 'D4', 'E4', 'F4']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN

    # 収入の部
    ws['A5'] = "【収入】"
    ws['A5'].font = HEADER_FONT

    income_items = [
        ("売上入金", [4500000, 5000000, 5500000], "現金売上+売掛回収"),
        ("その他収入", [100000, 0, 50000], "雑収入など"),
    ]

    row = 6
    for name, values, note in income_items:
        ws[f'A{row}'] = f"  {name}"
        for i, val in enumerate(values):
            col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].fill = INPUT_FILL
            ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=SUM(B{row}:D{row})'
        ws[f'E{row}'].fill = CALC_FILL
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'F{row}'] = note
        row += 1

    ws[f'A{row}'] = "収入合計"
    ws[f'A{row}'].font = TOTAL_FONT
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'=SUM({col}6:{col}{row-1})'
        ws[f'{col}{row}'].fill = TOTAL_FILL
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = TOTAL_FONT
    ws[f'E{row}'] = f'=SUM(E6:E{row-1})'
    ws[f'E{row}'].fill = TOTAL_FILL
    ws[f'E{row}'].number_format = '#,##0'
    ws[f'E{row}'].font = TOTAL_FONT
    row += 1

    # 支出の部
    ws[f'A{row}'] = "【支出】"
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    expense_items = [
        ("仕入支払", [1800000, 2000000, 2200000], "買掛金支払"),
        ("人件費", [800000, 800000, 800000], "給与・賞与"),
        ("家賃", [200000, 200000, 200000], "事務所・店舗"),
        ("広告費", [300000, 400000, 350000], "マーケティング"),
        ("その他経費", [200000, 180000, 220000], "諸経費"),
        ("融資返済", [300000, 300000, 300000], "元本+利息"),
        ("税金", [0, 0, 500000], "法人税等"),
    ]

    for name, values, note in expense_items:
        ws[f'A{row}'] = f"  {name}"
        for i, val in enumerate(values):
            col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].fill = INPUT_FILL
            ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=SUM(B{row}:D{row})'
        ws[f'E{row}'].fill = CALC_FILL
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'F{row}'] = note
        row += 1

    ws[f'A{row}'] = "支出合計"
    ws[f'A{row}'].font = TOTAL_FONT
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'=SUM({col}10:{col}{row-1})'
        ws[f'{col}{row}'].fill = TOTAL_FILL
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = TOTAL_FONT
    ws[f'E{row}'] = f'=SUM(E10:E{row-1})'
    ws[f'E{row}'].fill = TOTAL_FILL
    ws[f'E{row}'].number_format = '#,##0'
    ws[f'E{row}'].font = TOTAL_FONT
    row += 2

    # 資金収支
    ws[f'A{row}'] = "資金収支"
    ws[f'A{row}'].font = TOTAL_FONT
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'={col}8-{col}17'
        ws[f'{col}{row}'].fill = TOTAL_FILL
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = TOTAL_FONT
    ws[f'E{row}'] = f'=SUM(B{row}:D{row})'
    ws[f'E{row}'].fill = TOTAL_FILL
    ws[f'E{row}'].number_format = '#,##0'
    ws[f'E{row}'].font = TOTAL_FONT
    row += 1

    # 現金残高
    ws[f'A{row}'] = "期首現金残高"
    ws['B' + str(row)] = 3000000
    ws['B' + str(row)].fill = INPUT_FILL
    ws['B' + str(row)].number_format = '#,##0'
    ws['C' + str(row)] = f'=B{row+1}'
    ws['C' + str(row)].fill = CALC_FILL
    ws['C' + str(row)].number_format = '#,##0'
    ws['D' + str(row)] = f'=C{row+1}'
    ws['D' + str(row)].fill = CALC_FILL
    ws['D' + str(row)].number_format = '#,##0'
    row += 1

    ws[f'A{row}'] = "期末現金残高"
    ws[f'A{row}'].font = TOTAL_FONT
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'={col}{row-2}+{col}{row-1}'
        ws[f'{col}{row}'].fill = TOTAL_FILL
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = TOTAL_FONT

    # 列幅設定
    set_column_width(ws, 1, 25)
    for i in range(2, 6):
        set_column_width(ws, i, 15)
    set_column_width(ws, 6, 25)

    wb.save(os.path.join(OUTPUT_DIR, "資金繰り表.xlsx"))
    print("[OK] 資金繰り表.xlsx 作成完了")


def create_template_5():
    """5. 商品別損益表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品別損益表"

    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = "商品別損益表"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:F2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # ヘッダー
    headers = ["商品名", "売上高（円）", "売上原価（円）", "粗利（円）", "粗利率（%）", "構成比（%）"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    # 商品データ
    products = [
        ("商品A", 2000000, 800000),
        ("商品B", 1500000, 600000),
        ("商品C", 1000000, 500000),
        ("商品D", 500000, 200000),
    ]

    row = 5
    for name, sales, cost in products:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = sales
        ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = cost
        ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].number_format = '#,##0'
        ws[f'D{row}'] = f'=B{row}-C{row}'
        ws[f'D{row}'].fill = CALC_FILL
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'] = f'=D{row}/B{row}*100'
        ws[f'E{row}'].fill = CALC_FILL
        ws[f'E{row}'].number_format = '0.0'
        ws[f'F{row}'] = f'=B{row}/$B$9*100'
        ws[f'F{row}'].fill = CALC_FILL
        ws[f'F{row}'].number_format = '0.0'
        row += 1

    # 合計
    ws[f'A{row}'] = "合計"
    ws[f'A{row}'].font = TOTAL_FONT
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'=SUM({col}5:{col}{row-1})'
        ws[f'{col}{row}'].fill = TOTAL_FILL
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = TOTAL_FONT
    ws[f'E{row}'] = f'=D{row}/B{row}*100'
    ws[f'E{row}'].fill = CALC_FILL
    ws[f'E{row}'].number_format = '0.0'
    ws[f'F{row}'] = f'=SUM(F5:F{row-1})'
    ws[f'F{row}'].fill = CALC_FILL
    ws[f'F{row}'].number_format = '0.0'

    # 列幅設定
    set_column_width(ws, 1, 20)
    for i in range(2, 7):
        set_column_width(ws, i, 15)

    wb.save(os.path.join(OUTPUT_DIR, "商品別損益表.xlsx"))
    print("[OK] 商品別損益表.xlsx 作成完了")


def create_template_6():
    """6. 予算実績管理表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "予算実績管理表"

    # タイトル
    ws.merge_cells('A1:G1')
    ws['A1'] = "月次予算実績管理表"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:G2')
    ws['A2'] = "【使い方】青色のセルに予算・実績を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 対象月
    ws['A4'] = "対象月"
    ws['B4'] = "2025年1月"
    ws['B4'].fill = INPUT_FILL
    ws['B4'].alignment = CENTER_ALIGN

    # ヘッダー
    headers = ["項目", "予算（円）", "実績（円）", "差異（円）", "達成率（%）", "前年実績（円）", "前年比（%）"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    # データ
    items = [
        ("売上高", 5000000, 4800000, 4500000),
        ("売上原価", 2000000, 1920000, 1800000),
        ("売上総利益", "=B7-B8", "=C7-C8", "=F7-F8"),
        ("販管費", 1500000, 1450000, 1400000),
        ("営業利益", "=B9-B10", "=C9-C10", "=F9-F10"),
    ]

    row = 7
    for item, budget, actual, last_year in items:
        ws[f'A{row}'] = item

        # 予算
        if isinstance(budget, str):
            ws[f'B{row}'] = budget
            ws[f'B{row}'].fill = CALC_FILL
        else:
            ws[f'B{row}'] = budget
            ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'

        # 実績
        if isinstance(actual, str):
            ws[f'C{row}'] = actual
            ws[f'C{row}'].fill = CALC_FILL
        else:
            ws[f'C{row}'] = actual
            ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].number_format = '#,##0'

        # 差異
        ws[f'D{row}'] = f'=C{row}-B{row}'
        ws[f'D{row}'].fill = CALC_FILL
        ws[f'D{row}'].number_format = '#,##0'

        # 達成率
        ws[f'E{row}'] = f'=C{row}/B{row}*100'
        ws[f'E{row}'].fill = CALC_FILL
        ws[f'E{row}'].number_format = '0.0'

        # 前年実績
        if isinstance(last_year, str):
            ws[f'F{row}'] = last_year
            ws[f'F{row}'].fill = CALC_FILL
        else:
            ws[f'F{row}'] = last_year
            ws[f'F{row}'].fill = INPUT_FILL
        ws[f'F{row}'].number_format = '#,##0'

        # 前年比
        ws[f'G{row}'] = f'=C{row}/F{row}*100'
        ws[f'G{row}'].fill = CALC_FILL
        ws[f'G{row}'].number_format = '0.0'

        if "利益" in item:
            ws[f'A{row}'].font = TOTAL_FONT
        row += 1

    # 列幅設定
    set_column_width(ws, 1, 20)
    for i in range(2, 8):
        set_column_width(ws, i, 15)

    wb.save(os.path.join(OUTPUT_DIR, "予算実績管理表.xlsx"))
    print("[OK] 予算実績管理表.xlsx 作成完了")


def create_template_7():
    """7. 経営ダッシュボード"""
    wb = Workbook()
    ws = wb.active
    ws.title = "経営ダッシュボード"

    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = "経営ダッシュボード"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:F2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算され、KPIの健全性を一目で確認できます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 対象月
    ws['A4'] = "対象月"
    ws['B4'] = "2025年1月"
    ws['B4'].fill = INPUT_FILL
    ws['B4'].alignment = CENTER_ALIGN

    # 基本データ入力
    ws['A6'] = "【基本データ】"
    ws['A6'].font = HEADER_FONT

    basic_data = [
        ("売上高", 5000000, "月間売上"),
        ("売上原価", 2000000, "仕入・製造原価"),
        ("販管費", 1500000, "人件費+経費"),
        ("営業利益", "=B7-B8-B9", "本業の利益"),
        ("現金残高", 3000000, "手元資金"),
        ("売掛金", 1500000, "未回収売上"),
        ("在庫", 800000, "在庫金額"),
        ("買掛金", 600000, "未払仕入"),
        ("借入金残高", 3000000, "融資残高"),
        ("月次返済額", 300000, "元本+利息"),
    ]

    row = 7
    for name, value, note in basic_data:
        ws[f'A{row}'] = f"  {name}"
        if isinstance(value, str):
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = CALC_FILL
        else:
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = note
        row += 1

    # KPI
    ws['A19'] = "【主要KPI】"
    ws['A19'].font = HEADER_FONT

    ws['A20'] = "指標"
    ws['B20'] = "数値"
    ws['C20'] = "判定基準"
    ws['D20'] = "ステータス"
    for col in ['A20', 'B20', 'C20', 'D20']:
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER_ALIGN

    kpis = [
        ("粗利率", "=(B7-B8)/B7*100", "30%以上が望ましい", "=IF(B21>=30,\"✓良好\",\"要改善\")"),
        ("営業利益率", "=B10/B7*100", "10%以上が望ましい", "=IF(B22>=10,\"✓良好\",\"要改善\")"),
        ("運転資金", "=B11+B12-B13", "プラスであること", "=IF(B23>0,\"✓良好\",\"要注意\")"),
        ("借入返済比率", "=B15/B10", "50%以下が望ましい", "=IF(B24<=0.5,\"✓良好\",\"要改善\")"),
        ("現金月商比", "=B11/B7", "1.0以上が安全", "=IF(B25>=1,\"✓良好\",\"要注意\")"),
    ]

    row = 21
    for name, formula, standard, status in kpis:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = formula
        ws[f'B{row}'].fill = CALC_FILL
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = standard
        ws[f'D{row}'] = status
        ws[f'D{row}'].fill = CALC_FILL
        row += 1

    # 列幅設定
    set_column_width(ws, 1, 25)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 30)
    set_column_width(ws, 4, 15)

    wb.save(os.path.join(OUTPUT_DIR, "経営ダッシュボード.xlsx"))
    print("[OK] 経営ダッシュボード.xlsx 作成完了")


def create_template_8():
    """8. 損益分岐点計算シート"""
    wb = Workbook()
    ws = wb.active
    ws.title = "損益分岐点計算"

    # タイトル
    ws.merge_cells('A1:D1')
    ws['A1'] = "損益分岐点計算シート"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = CENTER_ALIGN

    # 使い方説明
    ws.merge_cells('A2:D2')
    ws['A2'] = "【使い方】青色のセルに数値を入力してください。緑色のセルは自動計算されます。"
    ws['A2'].font = Font(name='メイリオ', size=9, italic=True)
    ws['A2'].alignment = LEFT_ALIGN

    # 入力データ
    ws['A4'] = "【入力データ】"
    ws['A4'].font = HEADER_FONT

    inputs = [
        ("商品単価（円/個）", 5000, "1個あたりの販売価格"),
        ("変動費（円/個）", 2000, "1個あたりの仕入・製造原価"),
        ("固定費（円/月）", 1500000, "人件費+家賃+経費"),
    ]

    row = 5
    for name, value, note in inputs:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = INPUT_FILL
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = note
        row += 1

    # 計算結果
    ws['A10'] = "【計算結果】"
    ws['A10'].font = HEADER_FONT

    results = [
        ("限界利益（円/個）", "=B5-B6", "単価-変動費"),
        ("限界利益率（%）", "=B11/B5*100", "限界利益÷単価"),
        ("損益分岐点売上高（円）", "=B7/B12*100", "固定費÷限界利益率"),
        ("損益分岐点販売数（個）", "=B13/B5", "損益分岐点売上÷単価"),
        ("現在の売上高（円）", 6000000, "実際の売上"),
        ("現在の販売数（個）", "=B15/B5", "実際の販売数"),
        ("安全余裕率（%）", "=(B15-B13)/B15*100", "(売上-損益分岐点)÷売上"),
    ]

    row = 11
    for name, value, note in results:
        ws[f'A{row}'] = name
        if isinstance(value, str):
            ws[f'B{row}'] = value
            if row == 15:
                ws[f'B{row}'].fill = INPUT_FILL
            else:
                ws[f'B{row}'].fill = CALC_FILL
        else:
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = INPUT_FILL

        if row in [11, 13, 14, 15, 16]:
            ws[f'B{row}'].number_format = '#,##0'
        else:
            ws[f'B{row}'].number_format = '0.0'

        if row in [13, 14, 17]:
            ws[f'A{row}'].font = TOTAL_FONT
            ws[f'B{row}'].font = TOTAL_FONT
            if row == 13 or row == 14:
                ws[f'B{row}'].fill = TOTAL_FILL

        ws[f'C{row}'] = note
        row += 1

    # 判定
    ws['A20'] = "【判定】"
    ws['A20'].font = HEADER_FONT

    ws['A21'] = "安全余裕率の評価"
    ws['B21'] = "=IF(B17>=30,\"✓安全\",IF(B17>=10,\"注意\",\"危険\"))"
    ws['B21'].fill = CALC_FILL
    ws['C21'] = "30%以上：安全、10-30%：注意、10%未満：危険"

    # 列幅設定
    set_column_width(ws, 1, 30)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 40)

    wb.save(os.path.join(OUTPUT_DIR, "損益分岐点計算シート.xlsx"))
    print("[OK] 損益分岐点計算シート.xlsx 作成完了")


def main():
    """メイン処理"""
    print("=" * 60)
    print("Excelテンプレート作成開始")
    print("=" * 60)

    try:
        create_template_1()
        create_template_2()
        create_template_3()
        create_template_4()
        create_template_5()
        create_template_6()
        create_template_7()
        create_template_8()

        print("=" * 60)
        print("[OK] すべてのExcelテンプレートが正常に作成されました")
        print(f"保存先: {OUTPUT_DIR}")
        print("=" * 60)
    except Exception as e:
        print(f"[ERROR] エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
